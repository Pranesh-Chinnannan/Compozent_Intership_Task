from openpyxl import Workbook

# Two sets of lists
list1 = ["apple", "banana", "orange", "grape", "watermelon"]
list2 = ["carrot", "potato", "tomato", "spinach", "lettuce"]

# Create a new Workbook
wb = Workbook()

# Select the active worksheet
ws = wb.active

# Write the first list to Excel
for index, item in enumerate(list1, start=1):
    ws.cell(row=index, column=1, value=item)

# Calculate the starting row for the second list
start_row_list2 = len(list1) + 2  # Add 2 for some separation between lists

# Write the second list to Excel
for index, item in enumerate(list2, start=start_row_list2):
    ws.cell(row=index, column=1, value=item)

# Save the workbook
wb.save("output.xlsx")
